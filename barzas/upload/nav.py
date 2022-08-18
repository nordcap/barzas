# pyinstaller nav.py --onefile
# pyinstaller --onefile --windowed nav.py
# from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
import re

# стили для отображения таблиц
fontBold = Font(bold=True)
borderCell = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"),
                    right=Side(border_style="thin"))


def totalMove():
    sheet = wb.create_sheet(title="result")
    i = 2
    sheet.cell(row=1, column=1).value = "Объект"
    sheet.cell(row=1, column=1).font = fontBold
    sheet.cell(row=1, column=1).border = borderCell

    sheet.cell(row=1, column=2).value = "Смена"
    sheet.cell(row=1, column=2).font = fontBold
    sheet.cell(row=1, column=2).border = borderCell

    sheet.cell(row=1, column=3).value = "Зона погрузки"
    sheet.cell(row=1, column=3).font = fontBold
    sheet.cell(row=1, column=3).border = borderCell

    sheet.cell(row=1, column=4).value = "Зона разгрузки"
    sheet.cell(row=1, column=4).font = fontBold
    sheet.cell(row=1, column=4).border = borderCell

    sheet.cell(row=1, column=5).value = "Ходки"
    sheet.cell(row=1, column=5).font = fontBold
    sheet.cell(row=1, column=5).border = borderCell

    sheet.cell(row=1, column=6).value = "Расстояние"
    sheet.cell(row=1, column=6).font = fontBold
    sheet.cell(row=1, column=6).border = borderCell

    sheet.cell(row=1, column=7).value = "Сумма расстояний"
    sheet.cell(row=1, column=7).font = fontBold
    sheet.cell(row=1, column=7).border = borderCell

    dict_auto = dict()
    for key_tonar in dict_move:
        sheet.cell(row=i, column=1).value = key_tonar  # тонар
        sheet.cell(row=i, column=1).font = fontBold
        sheet.cell(row=i, column=1).border = borderCell
        dict_auto[key_tonar] = dict()

        for key_date in dict_move[key_tonar]:
            sheet.cell(row=i, column=2).value = key_date  # дата
            sheet.cell(row=i, column=2).font = fontBold
            list_zone = list(set(dict_move[key_tonar][key_date]))
            list_zone.sort()
            dict_auto[key_tonar][key_date] = dict()

            for zone in list_zone:
                count_zone = dict_move[key_tonar][key_date].count(zone)  # кол-во ходок
                sheet.cell(row=i, column=3).value = zone[0]  # зона погрузки
                sheet.cell(row=i, column=3).border = borderCell
                sheet.cell(row=i, column=4).value = zone[1]  # зона разгрузки
                sheet.cell(row=i, column=4).border = borderCell
                sheet.cell(row=i, column=5).value = count_zone  # кол-во совпадений
                sheet.cell(row=i, column=5).border = borderCell
                # расчет расстояний по справочной таблице
                if (zone in dict_distance):
                    sheet.cell(row=i, column=6).value = dict_distance[zone]
                    sheet.cell(row=i, column=6).border = borderCell
                    sheet.cell(row=i, column=7).value = dict_distance[zone] * count_zone
                    sheet.cell(row=i, column=7).border = borderCell
                    # формируем словарь в которой занесем информацию по ходкам и расстоянию
                    # если расстояние есть в словаре, то суммируем число ходок, соотв этому расстоянию
                    if ((dict_distance[zone] in dict_auto[key_tonar][key_date]) == True):
                        tmp = dict_auto[key_tonar][key_date][dict_distance[zone]]
                        dict_auto[key_tonar][key_date][dict_distance[zone]] = tmp + count_zone
                    else:
                        dict_auto[key_tonar][key_date][dict_distance[zone]] = count_zone
                i = i + 1

    return dict_auto


##################################################################################################################
def totalReestr():
    sheet = wb.create_sheet(title="СводныйРеестр")
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_distance) + 1)
    sheet.merge_cells(start_row=1, start_column=len(list_distance) + 2, end_row=2, end_column=len(list_distance) + 2)

    sheet.cell(row=1, column=1).value = "№№"
    sheet.cell(row=1, column=1).font = fontBold
    sheet.cell(row=1, column=1).border = borderCell

    sheet.cell(row=1, column=2).value = "Количество рейсов"
    sheet.cell(row=1, column=2).font = fontBold
    sheet.cell(row=1, column=2).border = borderCell

    sheet.cell(row=1, column=len(list_distance) + 2).value = "Итого рейсов"
    sheet.cell(row=1, column=len(list_distance) + 2).font = fontBold
    sheet.cell(row=1, column=len(list_distance) + 2).border = borderCell

    # формирование заголовков таблицы -перечень расстояний (столбцы)
    col = 2
    for cur_dist in list_distance:
        sheet.cell(row=2, column=col).value = cur_dist
        sheet.cell(row=2, column=col).font = Font(bold=True, italic=True)
        sheet.cell(row=2, column=col).border = borderCell
        col = col + 1

    # заполнение таблицы- подсчет рейсов в зависимости от расстояния (за все время)
    row = 3
    col = 2
    # новый словарь с накоплением итогов по столбцам
    dictSum = dict.fromkeys(list_distance, 0)
    dictSum["total"] = 0
    for key_tonar in dict_move:
        sheet.cell(row=row, column=1).value = key_tonar  # номер тонара
        sheet.cell(row=row, column=1).border = borderCell
        total_route = 0
        for distCol in list_distance:
            # получить список маршрутов, соответствующих дистанции dist. Может быть несколько рейсов с равным расстоянием
            # dict_distance=  {('ЭКГ №113', 'ДСУ №1'): 1.7}
            lst_route = []
            for route, distance in dict_distance.items():
                if distance == distCol:
                    lst_route.append(route)
            print("lst_route=", lst_route)
            # по найденным маршрутам ведем поиск по словарю dict_move
            #dict_move=  {'Howo №100': {'01.12.2021. Смена 1': [('Склад ДСУ1,2,3', 'Ж/Д склад')}}
            tonar_values = dict_move.get(key_tonar)
            lst = list(tonar_values.values())
            # преобразуем из списка список в список 1 порядка
            #arr_route =  [('Склад ДСУ1,2,3', 'Ж/Д склад'),.....]
            arr_route = [x for y in lst for x in y]
            print("arr_route=", arr_route)

            # находим  кол-во вхождений эл-ов lst_route в arr_route
            count = 0
            for item in lst_route:
                count = arr_route.count(item) + count
            sheet.cell(row=row, column=col).value = count
            sheet.cell(row=row, column=col).border = borderCell
            dictSum[distCol] = dictSum[distCol] + count #итого внизу таблицы
            col = col + 1
            total_route = count + total_route


        # итого рейсов
        sheet.cell(row=row, column=col).value = total_route
        sheet.cell(row=row, column=col).border = borderCell
        dictSum["total"] = dictSum["total"] + total_route
        row = row + 1  # переход на новую строку
        col = 2
    # итоговая строка
    sheet.cell(row=row, column=1).value = "Итого"
    sheet.cell(row=row, column=1).font = fontBold
    sheet.cell(row=row, column=1).border = borderCell
    for key in dictSum:
        sheet.cell(row=row, column=col).value = dictSum[key]
        sheet.cell(row=row, column=col).border = borderCell
        col = col + 1


##################################################################################################################
def load_distance():
    '''загрузка в словарь расстояния между объектами и вычисление множества расстояний'''
    wb = load_workbook(filename="ТаблицаРасстояний.xlsx")
    sheet = wb.active
    MAXROW = sheet.max_row
    MAXCOLUMN = sheet.max_column
    arrColumn = []
    arrRow = []
    # получаем список столбцов - пунктов разгрузки
    for column in range(2, MAXCOLUMN + 1):
        arrColumn.append(sheet.cell(row=2, column=column).value)

    # получаем список строк - пунктов погрузки
    for row in range(3, MAXROW + 1):
        arrRow.append(sheet.cell(row=row, column=1).value)

    # формируем таблицу-словарь, где ключи будут кортежами {('ЭКГ №113', 'ДСУ №1'): 1.7}
    dict_distance = dict()
    for i in range(3, MAXROW + 1):
        for j in range(2, MAXCOLUMN + 1):
            if sheet.cell(row=i, column=j).value == None:
                dict_distance[(arrRow[i - 3], arrColumn[j - 2])] = 0
            else:
                dict_distance[(arrRow[i - 3], arrColumn[j - 2])] = sheet.cell(row=i, column=j).value

    # список расстояний
    list_distance = list(set(list(dict_distance.values())))
    list_distance.sort()

    return (dict_distance, list_distance[1:])


###########################################################################################################

def reportTonar(tonar, dictTime):
    '''создание листа в Эксель по каждому тонару'''
    sheet = wb.create_sheet(title=tonar)
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(list_distance) + 1)
    sheet.merge_cells(start_row=1, start_column=len(list_distance) + 2, end_row=2, end_column=len(list_distance) + 2)

    sheet.cell(row=1, column=1).value = "Дата"
    sheet.cell(row=1, column=1).font = fontBold
    sheet.cell(row=1, column=1).border = borderCell

    sheet.cell(row=1, column=2).value = str(tonar) + " - Количество рейсов "
    sheet.cell(row=1, column=2).font = fontBold
    sheet.cell(row=1, column=2).border = borderCell

    sheet.cell(row=1, column=len(list_distance) + 2).value = "Итого рейсов"
    sheet.cell(row=1, column=len(list_distance) + 2).font = fontBold
    sheet.cell(row=1, column=len(list_distance) + 2).border = borderCell

    # формирование заголовков таблицы -перечень расстояний (столбцы)
    col = 2
    for cur_dist in list_distance:
        sheet.cell(row=2, column=col).value = cur_dist
        sheet.cell(row=2, column=col).font = Font(bold=True, italic=True)
        sheet.cell(row=2, column=col).border = borderCell
        col = col + 1

    # заполнение таблицы- подсчет рейсов в зависимости от расстояния (за все время)
    row = 3
    col = 2
    # новый словарь с накоплением итогов по столбцам
    dictSum = dict.fromkeys(list_distance, 0)
    dictSum["total"] = 0

    for keyTime in dictTime:
        sheet.cell(row=row, column=1).value = keyTime  # смена
        sheet.cell(row=row, column=1).border = borderCell
        total_route = 0
        for cur_dist in list_distance:
            if (cur_dist in dictTime[keyTime]):  # если в словаре найдены ходки
                dictSum[cur_dist] = dictSum[cur_dist] + dictTime[keyTime][cur_dist]
                sheet.cell(row=row, column=col).value = dictTime[keyTime][cur_dist]
                total_route = total_route + dictTime[keyTime][cur_dist]
            sheet.cell(row=row, column=col).border = borderCell
            col = col + 1
        sheet.cell(row=row, column=col).value = total_route
        sheet.cell(row=row, column=col).border = borderCell
        dictSum["total"] = dictSum["total"] + total_route
        row = row + 1  # переход на новую строку
        col = 2
    # итоговая строка
    sheet.cell(row=row, column=1).value = "Итого"
    sheet.cell(row=row, column=1).font = fontBold
    sheet.cell(row=row, column=1).border = borderCell
    for key in dictSum:
        sheet.cell(row=row, column=col).value = dictSum[key]
        sheet.cell(row=row, column=col).border = borderCell
        col = col + 1


###########################################################################################################
###########################################################################################################
###########################################################################################################
# MAIN


dict_distance, list_distance = load_distance()  # {('ЭКГ №113', 'ДСУ №1'): 1.7}
# print("dict_distance=", dict_distance)
# print("list_distance=", list_distance)

wb = load_workbook(filename="nav.xlsx")
sheet = wb.active

MAXROW = sheet.max_row

dict_move = dict()  # словарь с распарсенными данными вида {'Howo №100': {'01.12.2021. Смена 1': [('Склад ДСУ1,2,3', 'Ж/Д склад')}}
name_date = ""

for row in range(5, MAXROW):
    # 1 ячейка
    find_str = str(sheet.cell(row=row, column=1).value)
    # ищем строку вида Тонар №2694
    match_tonar = re.findall(r"(\w+)+\s№(\w+)", find_str)
    if len(match_tonar) > 0:
        name_tonar = match_tonar[0][0] + " №" + match_tonar[0][1]
        # если ключа нет в словаре, то добавляем и продолжаем цикл
        if ((name_tonar in dict_move) == False):
            dict_move[name_tonar] = dict()
            continue
    # ищем строку вида 02.10.2021. Смена 2
    match_date = re.findall(r"(\d{2}\.\d{2}\.\d{4}(\s|\S)*)", find_str)
    if len(match_date) > 0:
        name_date = match_date[0][0]
        dict_move[name_tonar][name_date] = []
        continue

    match_num = re.findall(r"\d{1,3}", find_str)
    if len(match_num) > 0:
        # 2 Ячейка -Зона погрузки
        zone_1 = str(sheet.cell(row=row, column=2).value)
        # 3 Ячейка -Зона разгрузки
        zone_2 = str(sheet.cell(row=row, column=3).value)
        print(name_tonar)
        dict_move[name_tonar][name_date].append((zone_1, zone_2))

a = totalMove()  # создаем 1 лист

# для каждого тонара создаем отдельные листы отчетов - данные берутся из листа result
for tonar in sorted(a.keys()):
    reportTonar(tonar, a[tonar])

# формируем отчет сводного реестра
totalReestr()

# формируем реестр путевых листов по каждому автомобилю
# reestrAuto()

wb.save(filename="nav_output.xlsx")
