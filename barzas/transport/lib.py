from openpyxl import load_workbook
from decimal import Decimal
import re
from .models import Distance


# функция обработки листа навигации и возвр. словарь ходок без учета кол-ва ходок
def calc_trip(sheet, MAXROW):
    # создаем словарь с распарсенными данными вида {'Howo №100': {'01.12.2021. Смена 1': [('Склад ДСУ1,2,3', 'Ж/Д склад'),...}}
    # global name_tonar
    dict_move = dict()
    name_date = ""

    for row in range(5, MAXROW):
        # 1 ячейка
        find_str = str(sheet.cell(row=row, column=1).value)
        # ищем строку вида Тонар №2694
        match_tonar = re.findall(r".*\s№.*", find_str)
        if len(match_tonar) > 0:
            name_tonar = match_tonar[0]
            # если ключа нет в словаре, то добавляем и продолжаем цикл
            if name_tonar not in dict_move:
                dict_move[name_tonar] = dict()
                continue
        # ищем строку вида 02.10.2021. Смена 2
        match_date = re.findall(r"(\d{2}\.\d{2}\.\d{4}(\s|\S)*)", find_str)
        if len(match_date) > 0:
            name_date = match_date[0][0]
            dict_move[name_tonar][name_date] = []
            continue
        # ищем цифру, которая находится на одной строке с искомыми данными
        match_num = re.findall(r"\d{1,3}", find_str)
        if len(match_num) > 0:
            # 2 Ячейка -Зона погрузки
            zone_1 = str(sheet.cell(row=row, column=2).value)
            # 3 Ячейка -Зона разгрузки
            zone_2 = str(sheet.cell(row=row, column=3).value)
            dict_move[name_tonar][name_date].append((zone_1, zone_2))
    return dict_move


# функция расширения calc_trip для добавления в словарь кол-ва ходок
# {'Foton Auman №О020НК142': {'01.06.2022. Смена 1': {('Склад ДСУ1,2,3', 'Ж/Д склад'): 1, ...}
def calc_trip_adv(dict_move):
    dict_auto = dict()
    for key_tonar in dict_move:
        dict_auto[key_tonar] = dict()
        for key_date in dict_move[key_tonar]:
            list_zone = list(set(dict_move[key_tonar][key_date]))
            list_zone.sort()
            dict_auto[key_tonar][key_date] = dict()
            for zone in list_zone:
                count_zone = dict_move[key_tonar][key_date].count(zone)  # кол-во ходок
                dict_auto[key_tonar][key_date][zone] = count_zone
    return dict_auto


def handle_upload_file(file):
    # основная функция обработки
    wb = load_workbook(filename=file)
    sheet = wb.active
    dict_move = calc_trip(sheet, sheet.max_row)
    # {'Howo №100': {'01.12.2021. Смена 1': [('Склад ДСУ1,2,3', 'Ж/Д склад')}}
    dict_trip = calc_trip_adv(dict_move)
    lst_distance = get_distance_between_store()  # [0.3, 0.7, 1.7, .....]
    dict_route = get_total_reestr(lst_distance, dict_move)

    return (dict_trip, dict_route, lst_distance)


def get_total_reestr(lst_distance, dict_move):
    # функция получение данных для сводного реестра ходок
    # {'Foton Auman №О020НК142':{1.7: 45, 3.0:2,}, ...}
    result = dict()
    result_with_list = dict()

    dict_distance = get_route_between_store()  # {('ЭКГ №113', 'ДСУ №1'): 1.7, .....}
    for key_tonar in dict_move:
        result[key_tonar] = dict()
        list_route = []

        for distCol in lst_distance:
            # получить список маршрутов, соответствующих дистанции dist.  Может быть несколько рейсов с равным расстоянием
            lst_route = []
            for route, distance in dict_distance.items():
                if distance == distCol:
                    lst_route.append(route)

            # из dict_move получаем все ходки в один список
            # dict_move=  {'Howo №100': {'01.12.2021. Смена 1': [('Склад ДСУ1,2,3', 'Ж/Д склад')}}
            tonar_values = dict_move.get(key_tonar)
            lst = list(tonar_values.values())
            arr_route = [x for y in lst for x in y]  # маршруты перевозки тонара

            # находим  кол-во вхождений эл-ов lst_route в arr_route
            count = 0
            for item in lst_route:
                count = arr_route.count(item) + count
            # записываем в результат
            result[key_tonar][distCol] = count
            list_route.append(count)
        total_sum = sum(result[key_tonar].values())  # колонка Всего по тонару
        list_route.append(total_sum)
        result_with_list[key_tonar] = list_route

    return result_with_list


def get_distance_between_store():
    # функция получение списка расстояний между объектами
    result = []
    queryset = Distance.objects.all().order_by('distance')
    # .distinct('distance') - не поддерживается в sqlite
    for obj in queryset:
        result.append(obj.distance)
    result = list(set(result))
    return sorted(result)


# получение словаря расстояний между объектами
def get_route_between_store():
    result = dict()
    queryset = Distance.objects.all()
    for obj in queryset:
        result[(obj.point_start.name, obj.point_final.name)] = obj.distance
    return result
